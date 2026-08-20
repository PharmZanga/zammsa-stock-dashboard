import json
import re
from pathlib import Path

import openpyxl


NEW_WORKBOOK = Path(
    r"C:\Users\Zanga Musakuzi\Desktop\zammsa folder\weekly inventory emms stock status\august\Stock position 07-14 August.xlsx"
)

NEW_SHEETS = [
    ("EMMS 07 JULY", "EMMS", "2026-08-07", "7 August 2026"),
    ("EMMS 14JULY", "EMMS", "2026-08-14", "14 August 2026"),
    ("LAB 7 JULY", "LAB", "2026-08-07", "7 August 2026"),
    ("LAB14 JULY", "LAB", "2026-08-14", "14 August 2026"),
]

OUTPUT_PATH = Path("src/weeklyAvailability.js")
MAX_DATA_ROWS = 2000


def clean_text(value):
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split())


def normalize_item(value):
    return re.sub(r"[^a-z0-9]+", " ", clean_text(value).lower()).strip()


def is_number(value):
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def is_availability(value):
    return is_number(value) and 0 <= float(value) <= 1


def find_overall(ws):
    for row in ws.iter_rows(
        min_row=1, max_row=min(ws.max_row or 1, MAX_DATA_ROWS), values_only=True
    ):
        cells = list(row)
        for index, value in enumerate(cells):
            if "overall" in clean_text(value).lower():
                for candidate in cells[index + 1 : index + 4]:
                    if is_availability(candidate):
                        return float(candidate)
    raise ValueError(f"Missing calculated overall availability in {ws.title}")


def find_category_columns(ws):
    for row in ws.iter_rows(
        min_row=1, max_row=min(ws.max_row or 1, MAX_DATA_ROWS), values_only=True
    ):
        cells = list(row)
        for index in range(len(cells) - 1):
            left = clean_text(cells[index]).lower()
            right = clean_text(cells[index + 1]).lower()
            if left == "category" and "availability" in right:
                return index, index + 1
            if left == "product category" and "availability" in right:
                return index, index + 1
    raise ValueError(f"Missing category columns in {ws.title}")


def extract_categories(ws, name_col, value_col):
    categories = []
    started = False
    blanks = 0
    for row in ws.iter_rows(
        min_row=1, max_row=min(ws.max_row or 1, MAX_DATA_ROWS), values_only=True
    ):
        cells = list(row) + [None] * max(0, value_col + 1 - len(row))
        name = clean_text(cells[name_col])
        value = cells[value_col]
        lower = name.lower()
        if lower in {"", "category", "product category", "overall stock percentage", "overall availability="}:
            if started:
                blanks += 1
                if blanks >= 8:
                    break
            continue
        if is_availability(value):
            categories.append([name, None, None, round(float(value), 6)])
            started = True
            blanks = 0
        elif started:
            blanks += 1
            if blanks >= 8:
                break
    return categories


def extract_items(ws, summary_name_col, category_names):
    items = []
    current_category = None
    category_lookup = {normalize_item(name): name for name in category_names}
    max_col = min(ws.max_column or 12, 16)

    for row in ws.iter_rows(
        min_row=1, max_row=min(ws.max_row or 1, MAX_DATA_ROWS), values_only=True
    ):
        cells = list(row) + [None] * max(0, max_col - len(row))
        for name_col in range(max_col - 1):
            if name_col == summary_name_col:
                continue
            name = clean_text(cells[name_col])
            value = cells[name_col + 1]
            if not name or not is_availability(value):
                continue

            normalized = normalize_item(name)
            if normalized in category_lookup:
                current_category = category_lookup[normalized]
                continue

            left = cells[name_col - 1] if name_col > 0 else None
            has_item_marker = is_number(left) or clean_text(left).startswith("#REF")
            if current_category and has_item_marker and float(value) in {0, 1}:
                items.append(
                    {
                        "item": name,
                        "category": current_category,
                        "available": int(value),
                    }
                )
                break
            if not has_item_marker and is_availability(value):
                current_category = name
    return items


def extract_sheet(ws, programme, date, label):
    overall = find_overall(ws)
    category_name_col, category_value_col = find_category_columns(ws)
    categories = extract_categories(ws, category_name_col, category_value_col)
    items = extract_items(ws, category_name_col, [category[0] for category in categories])

    category_totals = {}
    for item in items:
        values = category_totals.setdefault(item["category"], [0, 0])
        values[0] += item["available"]
        values[1] += 1

    categories = [
        [
            name,
            category_totals.get(name, [None, None])[0],
            category_totals.get(name, [None, None])[1],
            availability,
        ]
        for name, _, _, availability in categories
    ]

    total = len(items) if programme == "EMMS" else None
    available = sum(item["available"] for item in items) if programme == "EMMS" else None
    unavailable = total - available if programme == "EMMS" else None

    return {
        "date": date,
        "label": label,
        "programme": programme,
        "available": available,
        "total": total,
        "unavailable": unavailable,
        "availability": round(overall, 6),
        "categories": categories,
        "_items": items,
    }


def calculate_changes(reports, programme):
    programme_reports = sorted(
        [report for report in reports if report["programme"] == programme],
        key=lambda report: report["date"],
    )
    changes = []
    for previous, current in zip(programme_reports, programme_reports[1:]):
        previous_items = {normalize_item(item["item"]): item for item in previous["_items"]}
        current_items = {normalize_item(item["item"]): item for item in current["_items"]}
        matched = previous_items.keys() & current_items.keys()
        newly_unavailable = []
        recovered = []
        for key in matched:
            before = previous_items[key]
            after = current_items[key]
            detail = {"item": after["item"], "category": after["category"]}
            if before["available"] == 1 and after["available"] == 0:
                newly_unavailable.append(detail)
            elif before["available"] == 0 and after["available"] == 1:
                recovered.append(detail)
        changes.append(
            {
                "from": previous["label"],
                "to": current["label"],
                "newlyUnavailable": sorted(newly_unavailable, key=lambda item: item["item"]),
                "recovered": sorted(recovered, key=lambda item: item["item"]),
            }
        )
    return changes


def load_existing_output():
    text = OUTPUT_PATH.read_text(encoding="utf-8")
    match = re.search(r"export const weeklyAvailability = (\{.*\});\s*$", text, re.S)
    if not match:
        raise ValueError(f"Could not parse {OUTPUT_PATH}")
    return json.loads(match.group(1))


def main():
    output = load_existing_output()
    workbook = openpyxl.load_workbook(NEW_WORKBOOK, data_only=True, read_only=True)
    reports = [extract_sheet(workbook[sheet_name], programme, date, label) for sheet_name, programme, date, label in NEW_SHEETS]

    public_reports = [
        {key: value for key, value in report.items() if key != "_items"}
        for report in reports
    ]
    report_positions = {
        (report["programme"], report["date"]): index
        for index, report in enumerate(output["reports"])
    }
    for report in public_reports:
        key = (report["programme"], report["date"])
        if key in report_positions:
            output["reports"][report_positions[key]] = report
        else:
            output["reports"].append(report)

    changes_by_programme = {
        programme: calculate_changes(reports, programme) for programme in ("EMMS", "LAB")
    }
    for programme, changes in changes_by_programme.items():
        existing_changes = output["changesByProgramme"].setdefault(programme, [])
        change_positions = {
            (change["from"], change["to"]): index
            for index, change in enumerate(existing_changes)
        }
        for change in changes:
            key = (change["from"], change["to"])
            if key in change_positions:
                existing_changes[change_positions[key]] = change
            else:
                existing_changes.append(change)
    output["changes"] = output["changesByProgramme"]["EMMS"][-1]

    OUTPUT_PATH.write_text(
        "export const weeklyAvailability = " + json.dumps(output, indent=2) + ";\n",
        encoding="utf-8",
    )
    print(
        json.dumps(
            {
                "reports": [
                    {
                        "programme": report["programme"],
                        "date": report["date"],
                        "availability": report["availability"],
                        "categories": len(report["categories"]),
                        "items": len(report["_items"]),
                    }
                    for report in reports
                ],
                "latestChanges": {
                    programme: {
                        "from": changes[-1]["from"],
                        "to": changes[-1]["to"],
                        "newlyUnavailable": len(changes[-1]["newlyUnavailable"]),
                        "recovered": len(changes[-1]["recovered"]),
                    }
                    for programme, changes in changes_by_programme.items()
                },
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
